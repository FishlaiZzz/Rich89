import json
import sys
import openpyxl
import re

sys.stdout.reconfigure(encoding='utf-8')

js_path = r"D:\Antigravity\Rich\index.js"
xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
txt_89_path = r"D:\Antigravity\Rich\reference\第89期組別正副家長名單.txt"

# ----------------- 1. Update index.js to limit tags to max 3 -----------------
with open(js_path, "r", encoding="utf-8") as f:
    js_content = f.read()

# Extract groupData
start_marker = "  const groupData = "
end_marker = "};"

start_idx = js_content.find(start_marker)
if start_idx == -1:
    print("Could not find start of groupData")
    sys.exit(1)

start_json_idx = start_idx + len(start_marker)
end_idx = js_content.find(end_marker, start_json_idx)
if end_idx == -1:
    print("Could not find end of groupData")
    sys.exit(1)

json_str = js_content[start_json_idx:end_idx+1].strip()
group_data = json.loads(json_str)

truncated_in_js = 0
for group_id, group in group_data.items():
    for member in group["members"]:
        tags = member["tags"]
        if len(tags) > 3:
            print(f"JS: Truncating tags for '{member['name']}' from {len(tags)} to 3.")
            print(f"  Before: {tags}")
            member["tags"] = tags[:3]
            print(f"  After : {member['tags']}")
            truncated_in_js += 1

# Format back
formatted_json = json.dumps(group_data, indent=2, ensure_ascii=False)
indented_json = "\n".join("  " + line for line in formatted_json.splitlines())
new_js_content = js_content[:start_json_idx] + indented_json + ";\n" + js_content[end_idx+len(end_marker):]

with open(js_path, "w", encoding="utf-8") as f:
    f.write(new_js_content)

print(f"Updated index.js (truncated {truncated_in_js} profiles).")

# ----------------- 2. Update RAY哥-實戰班PPT2.xlsx to limit tags to max 3 -----------------
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

# Parse parent list first
with open(txt_89_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

parents = []
for line in lines:
    line = line.strip()
    if not line or line.startswith("【"):
        continue
    parts = line.split()
    match = re.match(r'^\d+(\S+)', parts[0])
    if match:
        p1 = match.group(1).strip()
    else:
        p1 = parts[0].strip()
    parents.append(p1)
    if len(parts) > 1:
        parents.append(parts[1].strip())

truncated_in_excel = 0
for r in range(2, sheet.max_row + 1):
    name = sheet.cell(row=r, column=1).value
    if name:
        name = name.strip()
        if name in parents:
            tags_raw = sheet.cell(row=r, column=3).value
            if tags_raw:
                tags = [t.strip() for t in tags_raw.split("\n") if t.strip()]
                if len(tags) > 3:
                    print(f"Excel: Truncating tags for '{name}' in Row {r} from {len(tags)} to 3.")
                    truncated_tags = tags[:3]
                    new_tags_str = "\n".join(truncated_tags)
                    sheet.cell(row=r, column=3).value = new_tags_str
                    sheet.cell(row=r, column=4).value = "已更新"
                    truncated_in_excel += 1

wb.save(xlsx_path)
print(f"Updated Excel file (truncated {truncated_in_excel} profiles).")
