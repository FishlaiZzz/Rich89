import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

missing = ['楊琬琳', '王佩婷', '楊媛媛', '王宥鈞', '謝彥馨', '琬琳', '佩婷', '媛媛', '宥鈞', '彥馨']

xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT.xlsx"
wb = openpyxl.load_workbook(xlsx_path)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\nSearching sheet: {sheetname}")
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val:
                val_str = str(val)
                for m in missing:
                    if m in val_str:
                        print(f"  Found '{m}' at Row {r}, Col {c}: '{val_str}'")
