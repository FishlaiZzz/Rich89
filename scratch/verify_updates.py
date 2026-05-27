import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

print("Rows with '已更新' in Column D:")
found_any = False
for r in range(2, sheet.max_row + 1):
    d_val = sheet.cell(row=r, column=4).value
    if d_val == "已更新":
        name = sheet.cell(row=r, column=1).value
        tags = sheet.cell(row=r, column=3).value
        print(f"Row {r:3d}: Name='{name}', Tags:\n{tags}\n")
        found_any = True

if not found_any:
    print("None found!")
