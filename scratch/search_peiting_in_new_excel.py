import openpyxl

xlsx_path = r"d:\Antigravity\Rich\reference\Rich正副家長.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
sheet = wb["個人資料"]

print("Searching for row containing '王佩婷' or '佩婷' or '0':")
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 4)]
    if any(vals) and any("王佩婷" in str(v) or "佩婷" in str(v) for v in vals if v is not None):
        print(f"Row {r}: {vals}")
