import re
import os
import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

# Paths
xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
txt_89_path = r"D:\Antigravity\Rich\reference\第89期組別正副家長名單.txt"
js_path = r"D:\Antigravity\Rich\index.js"

# 1. Parse parent list
with open(txt_89_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

parents = []
for line in lines:
    line = line.strip()
    if not line or line.startswith("【"):
        continue
    parts = line.split()
    match = re.match(r'^\d+(\S+)', parts[0])
    if match:
        p1 = match.group(1).strip()
    else:
        p1 = parts[0].strip()
    parents.append(p1)
    if len(parts) > 1:
        parents.append(parts[1].strip())

print(f"Loaded {len(parents)} parents.")

# 2. Parse Excel
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

excel_dict = {}
for r in range(2, sheet.max_row + 1):
    name = sheet.cell(row=r, column=1).value
    if name:
        name = name.strip()
        talent = sheet.cell(row=r, column=2).value
        if talent:
            talent = talent.strip()
        tags_raw = sheet.cell(row=r, column=3).value
        tags = []
        if tags_raw:
            tags = [t.strip() for t in tags_raw.split("\n") if t.strip()]
        excel_dict[name] = {"talent": talent, "tags": tags}

# 3. Load index.js
with open(js_path, "r", encoding="utf-8") as f:
    js_content = f.read()

# Extract groupData
start_marker = "  const groupData = "
end_marker = "};"

start_idx = js_content.find(start_marker)
if start_idx == -1:
    print("Could not find start of groupData")
    sys.exit(1)

start_json_idx = start_idx + len(start_marker)
end_idx = js_content.find(end_marker, start_json_idx)
if end_idx == -1:
    print("Could not find end of groupData")
    sys.exit(1)

json_str = js_content[start_json_idx:end_idx+1].strip()
group_data = json.loads(json_str)

# 4. Update group_data based on Excel
updated_count = 0
for group_id, group in group_data.items():
    for member in group["members"]:
        name = member["name"].strip()
        if name in excel_dict:
            excel_info = excel_dict[name]
            
            # Print if changing
            if member["talent"] != excel_info["talent"] or member["tags"] != excel_info["tags"]:
                print(f"Updating '{name}' in Group {group_id}:")
                print(f"  Talent: '{member['talent']}' -> '{excel_info['talent']}'")
                print(f"  Tags: {member['tags']} -> {excel_info['tags']}")
                member["talent"] = excel_info["talent"]
                member["tags"] = excel_info["tags"]
                updated_count += 1

# 5. Format group_data back to JS string
# We want to format it beautifully with 2 spaces indent, and 2 spaces leading indent on every line
formatted_json = json.dumps(group_data, indent=2, ensure_ascii=False)
# Align indentation (prepend 2 spaces to each line)
indented_json = "\n".join("  " + line for line in formatted_json.splitlines())

# Remove original json and insert the new one
new_js_content = js_content[:start_json_idx] + indented_json + ";\n" + js_content[end_idx+len(end_marker):]

with open(js_path, "w", encoding="utf-8") as f:
    f.write(new_js_content)

print(f"\nSuccessfully updated {updated_count} parent profiles in {js_path}!")
