import openpyxl

xlsx_path = r"d:\Antigravity\Rich\reference\Rich正副家長.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

results = []

for sname in wb.sheetnames:
    sheet = wb[sname]
    results.append(f"\n================ SHEET: {sname} (max_row={sheet.max_row}, max_col={sheet.max_column}) ================")
    
    # Read first 15 rows and 10 columns
    for r in range(1, min(sheet.max_row + 1, 30)):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 10))]
        results.append(f"Row {r}: {row_vals}")

with open(r"scratch\inspect_rich_parents_results.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(results))

print("Inspection completed.")
