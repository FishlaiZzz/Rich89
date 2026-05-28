import openpyxl
import os

files = [f for f in os.listdir("reference") if f.endswith(".xlsx")]
print("Excel files in reference:", files)

results = []

for fname in files:
    fpath = os.path.join("reference", fname)
    results.append(f"\n================ FILE: {fname} ================")
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for r in range(1, sheet.max_row + 1):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                row_str = " | ".join([str(x) for x in row_vals if x is not None])
                
                # Check for keywords
                for kw in ["媛媛", "琬琳", "彥馨", "佩婷"]:
                    if kw in row_str:
                        results.append(f"  [{sname}] Row {r} ({kw}): {row_str}")
    except Exception as e:
        results.append(f"Error reading {fname}: {e}")

with open(r"scratch\search_missing_parents_excel.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(results))

print("Completed search for missing parents in excel.")
