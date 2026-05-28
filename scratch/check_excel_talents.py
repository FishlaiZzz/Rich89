import openpyxl

wb = openpyxl.load_workbook(r"reference\RAY哥-實戰班PPT2.xlsx", data_only=True)
sheet = wb["Rich89"]

names = ["楊琬琳", "楊媛媛", "王宥鈞", "謝彥馨"]
for r in range(1, sheet.max_row + 1):
    val_name = sheet.cell(row=r, column=1).value
    if val_name in names:
        val_talent = sheet.cell(row=r, column=2).value
        print(f"Name: {val_name} | Excel Talent: {val_talent}")
