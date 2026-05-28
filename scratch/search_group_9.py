import openpyxl

wb = openpyxl.load_workbook(r"reference\RAY哥-實戰班PPT2.xlsx", data_only=True)
sheet = wb["Rich89"]

for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 5)]
    if any(vals) and any("皮蛙" in str(v) for v in vals if v is not None):
        print(f"Row {r}: {vals}")
