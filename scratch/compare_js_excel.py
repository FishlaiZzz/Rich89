import re
import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# 1. Parse parent list
txt_89_path = r"D:\Antigravity\Rich\reference\第89期組別正副家長名單.txt"
with open(txt_89_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

parents = []
for line in lines:
    line = line.strip()
    if not line or line.startswith("【"):
        continue
    parts = line.split()
    match = re.match(r'^\d+(\S+)', parts[0])
    if match:
        p1 = match.group(1).strip()
    else:
        p1 = parts[0].strip()
    parents.append(p1)
    if len(parts) > 1:
        parents.append(parts[1].strip())

# 2. Parse index.js groupData
js_path = r"D:\Antigravity\Rich\index.js"
with open(js_path, "r", encoding="utf-8") as f:
    js_content = f.read()

# Parse members in js
js_members = re.findall(r'"?name"?:\s*"([^"]+)"[^{}]+"?talent"?:\s*"([^"]+)"[^{}]+"?role"?:\s*"([^"]+)"[^{}]+"?image"?:\s*"([^"]+)"[^{}]+"?tags"?:\s*\[([^\]]+)\]', js_content)

js_dict = {}
for m in js_members:
    name = m[0].strip()
    talent = m[1].strip()
    role = m[2].strip()
    image = m[3].strip()
    tags_raw = m[4]
    tags = [t.replace('"', '').replace("'", "").replace("＃", "#").strip() for t in tags_raw.split(",")]
    js_dict[name] = {"talent": talent, "role": role, "image": image, "tags": tags}

# 3. Parse Excel
xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

excel_dict = {}
for r in range(2, sheet.max_row + 1):
    name = sheet.cell(row=r, column=1).value
    if name:
        name = name.strip()
        talent = sheet.cell(row=r, column=2).value
        if talent:
            talent = talent.strip()
        tags_raw = sheet.cell(row=r, column=3).value
        tags = []
        if tags_raw:
            tags = [t.strip().replace("＃", "#") for t in tags_raw.split("\n") if t.strip()]
        excel_dict[name] = {"talent": talent, "tags": tags}

# 4. Compare
print("================ DIFFERENCE REPORT (index.js vs Excel) ================")
for parent in parents:
    print(f"\nParent: '{parent}'")
    if parent not in js_dict:
        print("  - NOT IN index.js (ERROR!)")
        continue
    
    js_data = js_dict[parent]
    
    if parent not in excel_dict:
        print("  - NOT IN RAY哥-實戰班PPT2.xlsx")
        continue
        
    excel_data = excel_dict[parent]
    
    # Compare talent
    if js_data["talent"] != excel_data["talent"]:
        print(f"  - Talent Difference: JS='{js_data['talent']}', Excel='{excel_data['talent']}'")
        
    # Compare tags
    js_tags = [t.replace("#", "") for t in js_data["tags"]]
    excel_tags = [t.replace("#", "") for t in excel_data["tags"]]
    if js_tags != excel_tags:
        print(f"  - Tags Difference:")
        print(f"    JS   : {js_tags}")
        print(f"    Excel: {excel_tags}")
