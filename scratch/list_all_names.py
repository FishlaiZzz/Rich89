import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

print("All names in Column A of Rich89 sheet:")
names = []
for r in range(2, sheet.max_row + 1):
    val = sheet.cell(row=r, column=1).value
    if val:
        names.append(val.strip())

print(names)
