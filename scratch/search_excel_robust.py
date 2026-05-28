import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

files = [f for f in os.listdir("reference") if f.endswith(".xlsx")]
print("Excel files in reference:", files)

for fname in files:
    fpath = os.path.join("reference", fname)
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val and ("佩婷" in str(val) or "王佩婷" in str(val)):
                        print(f"FOUND: [{fname}] -> Sheet: [{sname}] -> Row {r}, Col {c}: {val}")
    except Exception as e:
        print(f"Error reading {fname}: {e}")
