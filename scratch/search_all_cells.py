import openpyxl
import os

files = [f for f in os.listdir("reference") if f.endswith(".xlsx")]
print("Excel files in reference:", files)

results = []

for fname in files:
    fpath = os.path.join("reference", fname)
    results.append(f"\n================ FILE: {fname} ================")
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            results.append(f"\n--- Sheet: {sname} (max_row={sheet.max_row}, max_col={sheet.max_column}) ---")
            
            # Print column headers or first few rows to see what columns exist
            for r in range(1, min(sheet.max_row + 1, 10)):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 15))]
                results.append(f"Row {r}: {row_vals}")
                
            # Search for ANY cell containing "佩婷" or "王" or "支持者"
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val and ("佩婷" in str(val) or "王" in str(val) or "支持" in str(val)):
                        # If "佩婷" matches, print it!
                        if "佩婷" in str(val):
                            results.append(f"  [MATCH PEITING] Row {r}, Col {c} (Value): {val}")
                        # If it matches "王" or "支持者", let's see if it's near Group 9
                        # We can just look at row context if it's "王" and "支持者"
                        if "王" in str(val) and "支持" in str(val):
                            results.append(f"  [MATCH WANG+SUPPORT] Row {r}, Col {c} (Value): {val}")
    except Exception as e:
        results.append(f"Error reading {fname}: {e}")

with open(r"scratch\search_all_cells.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(results))

print("Completed deep cell search.")
