import openpyxl
import os

files = [f for f in os.listdir("reference") if f.endswith(".xlsx")]
print("Excel files:", files)

results = []

for fname in files:
    fpath = os.path.join("reference", fname)
    results.append(f"\n--- Reading file: {fname} ---")
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            results.append(f"  Sheet: {sname} (max_row={sheet.max_row}, max_col={sheet.max_column})")
            
            # Let's search for "佩" or "婷" or "王" or look at rows that mention group 9 or 9組 or "支持者"
            for r in range(1, min(sheet.max_row + 1, 1000)):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 30))]
                row_str = " | ".join([str(x) for x in row_vals if x is not None])
                
                # Search for keywords
                if "佩婷" in row_str or "王佩婷" in row_str:
                    results.append(f"    [MATCH PEITING] Row {r}: {row_str}")
                elif "9" in row_str and ("支持" in row_str or "楊振達" in row_str):
                    results.append(f"    [MATCH G9/YANG] Row {r}: {row_str}")
    except Exception as e:
        results.append(f"  Error: {e}")

with open(r"scratch\excel_search_results.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(results))

print("Done. Written to scratch/excel_search_results.txt")
