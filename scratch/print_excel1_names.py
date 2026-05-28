import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

fname = 'RAY哥-實戰班PPT.xlsx'
fpath = os.path.join("reference", fname)

try:
    wb = openpyxl.load_workbook(fpath, data_only=True)
    for sname in wb.sheetnames:
        sheet = wb[sname]
        print(f"\nScanning sheet [{sname}] in [{fname}] (Total Rows: {sheet.max_row})")
        for r in range(1, sheet.max_row + 1):
            name = sheet.cell(row=r, column=1).value
            talent = sheet.cell(row=r, column=2).value
            tags = sheet.cell(row=r, column=3).value
            if name:
                name_str = str(name).strip()
                if "佩" in name_str or "婷" in name_str or "王" in name_str or "佩婷" in name_str:
                    print(f"Row {r:3d}: Name='{name_str}', Talent='{talent}', Tags='{tags}'")
except Exception as e:
    print(f"Error reading {fname}: {e}")
