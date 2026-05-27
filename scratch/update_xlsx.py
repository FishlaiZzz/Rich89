import re
import os
import sys
import openpyxl
import shutil

sys.stdout.reconfigure(encoding='utf-8')

# Paths
xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
backup_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2_backup.xlsx"
txt_89_path = r"D:\Antigravity\Rich\reference\第89期組別正副家長名單.txt"
tags_path = r"D:\Antigravity\Rich\reference\正副家長姓名標籤.txt"

# 1. Create backup of Excel
if not os.path.exists(backup_path):
    shutil.copy(xlsx_path, backup_path)
    print(f"Backed up original file to: {backup_path}")
else:
    print(f"Backup already exists at: {backup_path}")

# 2. Parse parent list
with open(txt_89_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

parents = []
for line in lines:
    line = line.strip()
    if not line or line.startswith("【"):
        continue
    parts = line.split()
    match = re.match(r'^\d+(\S+)', parts[0])
    if match:
        p1 = match.group(1).strip()
    else:
        p1 = parts[0].strip()
    parents.append(p1)
    if len(parts) > 1:
        parents.append(parts[1].strip())

print(f"Loaded {len(parents)} parents from list.")

# 3. Parse tags file
with open(tags_path, "r", encoding="utf-8") as f:
    tags_raw = f.read()

blocks = []
current_header = None
current_tags = []

lines = tags_raw.splitlines()
for line in lines:
    line = line.strip()
    if not line:
        continue
    if line.startswith("#") or line.startswith("＃"):
        if current_header:
            current_tags.append(line)
    else:
        if current_header and current_tags:
            blocks.append((current_header, current_tags))
        current_header = line
        current_tags = []

if current_header and current_tags:
    blocks.append((current_header, current_tags))

print(f"Loaded {len(blocks)} tag blocks from tag file.")

# Define mapping rules for parents to tag blocks
def get_matched_tags(parent):
    # Check manual/explicit rules first
    if parent == "林苡姗":
        for h, t in blocks:
            if "玟姍" in h:
                return t
    if parent == "徐瑩瑩":
        for h, t in blocks:
            if "聊癒瑩子" in h or "瑩子" in h:
                return t
                
    # Direct substring
    for h, t in blocks:
        if parent in h:
            return t
            
    # Substring of last 2 characters
    if len(parent) >= 3:
        short_name = parent[1:]
        for h, t in blocks:
            if short_name in h:
                return t
                
    return None

# Load Excel
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

print("\n--- Processing and Matching ---")
updated_count = 0
for row_idx in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row_idx, column=1).value
    if not name:
        continue
    name = name.strip()
    
    # Check if this person is in the 89th batch parent list
    if name not in parents:
        continue
        
    # Find updated tags for this parent
    updated_tags_list = get_matched_tags(name)
    if updated_tags_list:
        new_tags_str = "\n".join(updated_tags_list)
        old_tags_str = sheet.cell(row=row_idx, column=3).value
        
        # Normalize comparison (stripping extra newlines, space, etc.)
        norm_old = (old_tags_str or "").strip().replace("\r\n", "\n").replace("＃", "#")
        norm_new = new_tags_str.strip().replace("\r\n", "\n").replace("＃", "#")
        
        if norm_old != norm_new:
            print(f"Row {row_idx}: Updating '{name}'")
            print(f"  Old Tags:\n{old_tags_str}")
            print(f"  New Tags:\n{new_tags_str}")
            
            sheet.cell(row=row_idx, column=3).value = new_tags_str
            sheet.cell(row=row_idx, column=4).value = "已更新"
            updated_count += 1
        else:
            # If tags are identical, should we still mark as "已更新" if they had a match?
            # Let's see: "是否有更新標籤, 把標籤更新回 RAY哥-實戰班PPT2.xlsx 的 Rich 89 工作表中, 並在D欄標示"已更新""
            # Usually, marking it "已更新" indicates that this row belongs to the matched set.
            # Let's write "已更新" anyway or if it matched. Let's write "已更新" for all matched parents.
            print(f"Row {row_idx}: '{name}' tags are already up-to-date, marking as '已更新'.")
            sheet.cell(row=row_idx, column=4).value = "已更新"
            updated_count += 1

wb.save(xlsx_path)
print(f"\nDone! Saved {updated_count} rows marked as '已更新' in {xlsx_path}.")
