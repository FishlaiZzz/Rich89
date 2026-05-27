import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

missing = ['楊琬琳', '王佩婷', '楊媛媛', '王宥鈞', '謝彥馨']

xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

excel_names = []
for r in range(2, sheet.max_row + 1):
    val = sheet.cell(row=r, column=1).value
    if val:
        excel_names.append((r, val.strip()))

print("Fuzzy search in Excel names:")
for m in missing:
    found = False
    for r, name in excel_names:
        if m in name or name in m or m[1:] in name or name[1:] in m:
            print(f"  Missing '{m}' could be Row {r}: '{name}'")
            found = True
    if not found:
        print(f"  Missing '{m}': No similar name found.")
