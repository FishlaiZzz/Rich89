import openpyxl

wb = openpyxl.load_workbook(r"reference\RAY哥-實戰班PPT2.xlsx", data_only=True)
sheet = wb["Rich89"]

names = set()
for r in range(2, sheet.max_row + 1):
    val = sheet.cell(row=r, column=1).value
    if val:
        names.add(str(val).strip())

print(f"Total unique names in Rich89 sheet: {len(names)}")
with open(r"scratch\print_all_excel_names.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(sorted(list(names))))

print("Saved to scratch/print_all_excel_names.txt")
