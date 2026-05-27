import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_path = r"D:\Antigravity\Rich\reference\RAY哥-實戰班PPT2.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb['Rich89']

with open("scratch/excel_rows.txt", "w", encoding="utf-8") as f:
    f.write(f"Row, Name, Talent, Tags, D_Val\n")
    for r in range(1, sheet.max_row + 1):
        name = sheet.cell(row=r, column=1).value
        talent = sheet.cell(row=r, column=2).value
        tags = sheet.cell(row=r, column=3).value
        d_val = sheet.cell(row=r, column=4).value
        tags_repr = repr(tags) if tags else "None"
        f.write(f"{r}, {name}, {talent}, {tags_repr}, {d_val}\n")

print("Dumped all rows to scratch/excel_rows.txt")
