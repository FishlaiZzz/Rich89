import os
import shutil
import openpyxl
import re

ref_dir = r"d:\Antigravity\Rich\reference"
photo_dir = r"d:\Antigravity\Rich\reference\photo"
js_path = r"d:\Antigravity\Rich\index.js"
xlsx_path = r"d:\Antigravity\Rich\reference\Rich正副家長.xlsx"

print("================ MOVE AVATARS AND UPDATE TAGS ================")

# 1. Move avatar_*.png files to reference/photo/
avatars = [f for f in os.listdir(ref_dir) if f.startswith("avatar_") and f.endswith(".png")]
print(f"Found avatars to move: {avatars}")

for av in avatars:
    src = os.path.join(ref_dir, av)
    dest = os.path.join(photo_dir, av)
    try:
        shutil.move(src, dest)
        print(f"Moved {av} -> {dest}")
    except Exception as e:
        print(f"Error moving {av}: {e}")

# 2. Read latest tags from Rich正副家長.xlsx sheet "個人資料"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
sheet = wb["個人資料"]

excel_data = {}
for r in range(2, sheet.max_row + 1):
    name_val = sheet.cell(row=r, column=2).value
    if name_val is not None:
        name = str(name_val).strip()
        tags_val = sheet.cell(row=r, column=3).value
        tags_str = str(tags_val) if tags_val is not None else ""
        
        # Split tags by newline
        raw_tags = [t.strip() for t in tags_str.split("\n") if t.strip()]
        # Check if the Excel cell has literally '0' or is empty
        if len(raw_tags) == 1 and raw_tags[0] == "0":
            excel_data[name] = []
        else:
            excel_data[name] = raw_tags

# 3. Process index.js
with open(js_path, "r", encoding="utf-8") as f:
    js_content = f.read()

# Parent names for reference
parent_names = [
    "王慈惠", "蕭惠允", "許蓓瑾", "楊家瑜", "劉錦芳", "楊琬琳",
    "楊君慧", "周詩容", "黃雲嵩", "賴貞如", "王雅琳", "林苡姗",
    "陳瑩臻", "許淳皓", "林胤呈", "張庭熒", "楊振達", "王佩婷",
    "吳家怡", "徐瑩瑩", "江勁霖", "楊媛媛", "王宥鈞", "陳彥萍",
    "謝彥馨", "張家榮"
]

# We need to map which default avatar each person was originally using
# Female 1, Female 2, Male 1, Male 2.
# Let's map based on gender/profile name
avatar_fallbacks = {
    # Females using female_1
    "王慈惠": "avatar_female_1.png",
    "許蓓瑾": "avatar_female_1.png",
    "劉錦芳": "avatar_female_1.png",
    "楊君慧": "avatar_female_1.png",
    "賴貞如": "avatar_female_1.png",
    "林苡姗": "avatar_female_1.png",
    "張庭熒": "avatar_female_1.png",
    "吳家怡": "avatar_female_1.png",
    "楊媛媛": "avatar_female_1.png",
    "謝彥馨": "avatar_female_1.png",
    
    # Females using female_2
    "蕭惠允": "avatar_female_2.png",
    "楊家瑜": "avatar_female_2.png",
    "楊琬琳": "avatar_female_2.png",
    "周詩容": "avatar_female_2.png",
    "王雅琳": "avatar_female_2.png",
    "陳瑩臻": "avatar_female_2.png",
    "王佩婷": "avatar_female_2.png",
    "徐瑩瑩": "avatar_female_2.png",
    "陳彥萍": "avatar_female_2.png",
    
    # Males using male_1
    "黃雲嵩": "avatar_male_1.png",
    "林胤呈": "avatar_male_1.png",
    "江勁霖": "avatar_male_1.png",
    "張家榮": "avatar_male_1.png",
    
    # Males using male_2
    "許淳皓": "avatar_male_2.png",
    "楊振達": "avatar_male_2.png",
    "王宥鈞": "avatar_male_2.png",
}

photos = os.listdir(photo_dir)

log_lines = []

# We parse each profile block using regex and update:
# 1. image path (if no real photo, fallback to reference/photo/avatar_...)
# 2. tags list (pad with "(待補)" to exactly 3 tags)

# Regex to match each member's block in index.js:
# {
#   "name": "NAME",
#   "talent": "TALENT",
#   "role": "ROLE",
#   "image": "IMAGE",
#   "tags": [ ... ]
# }
member_block_pattern = r'\{\s*"name":\s*"([^"]+)",\s*"talent":\s*"([^"]+)",\s*"role":\s*"([^"]+)",\s*"image":\s*"([^"]+)",\s*"tags":\s*\[([^\]]+)\]'

def replace_member(match):
    name = match.group(1)
    talent = match.group(2)
    role = match.group(3)
    old_image = match.group(4)
    old_tags_raw = match.group(5)
    
    # --- A. Determine Image Path ---
    # Find matching photo files on disk (excluding avatar_ prefixes)
    real_photo_files = [p for p in photos if name in p and not p.startswith("avatar_")]
    if real_photo_files:
        new_image = f"reference/photo/{real_photo_files[0]}"
        # Update talent from photo file if needed
        photo_talent = real_photo_files[0].split("-")[0]
        new_talent = photo_talent
    else:
        # Fallback to avatar in reference/photo/
        fallback_file = avatar_fallbacks.get(name, "avatar_female_1.png")
        new_image = f"reference/photo/{fallback_file}"
        new_talent = talent # keep original talent if using avatar fallback
    
    # --- B. Determine Tags List ---
    # Find tag in excel mapping
    tags = []
    if name in excel_data:
        tags = excel_data[name]
    else:
        # Try substring match
        matched_name = None
        for ex_name in excel_data.keys():
            if name in ex_name or ex_name in name:
                matched_name = ex_name
                break
        if matched_name:
            tags = excel_data[matched_name]
        else:
            tags = []
            
    # Clean tags list
    tags = [t.strip() for t in tags if t.strip()]
    
    # Pad to exactly 3 tags with "(待補)"
    while len(tags) < 3:
        tags.append("(待補)")
    # Limit to maximum 3 tags just in case
    tags = tags[:3]
    
    # Format tags list string in JS format
    js_tags_str = '[\n            ' + ',\n            '.join([f'"{t}"' for t in tags]) + '\n          ]'
    
    log_lines.append(f"Member: {name} | Image: {new_image} | Talent: {new_talent} | Tags: {tags}")
    
    # Reconstruct the member block
    return f'{{\n          "name": "{name}",\n          "talent": "{new_talent}",\n          "role": "{role}",\n          "image": "{new_image}",\n          "tags": {js_tags_str}'

new_js_content = re.sub(member_block_pattern, replace_member, js_content, flags=re.DOTALL)

# Save the updated index.js
with open(js_path, "w", encoding="utf-8") as f:
    f.write(new_js_content)

# Save log
with open(r"d:\Antigravity\Rich\scratch\move_and_update_log.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines))

print("Successfully updated index.js and moved avatars! Log written to scratch/move_and_update_log.txt")
