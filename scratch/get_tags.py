import openpyxl
import json
import re

xlsx_path = r"d:\Antigravity\Rich\reference\Rich正副家長.xlsx"
txt_path = r"d:\Antigravity\Rich\reference\第89期組別正副家長名單.txt"

# 1. Parse parent names from the txt file
with open(txt_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

parent_names = []
for line in lines:
    line = line.strip()
    if not line or line.startswith("【"):
        continue
    match = re.match(r'^(\d+)\s*(\S+)\s+(\S+)', line)
    if match:
        parent_names.append(match.group(2).strip())
        parent_names.append(match.group(3).strip())

print(f"Parsed {len(parent_names)} parent names from txt file.")

# 2. Read Rich正副家長.xlsx
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
sheet = wb["個人資料"]

# Let's map all names in the excel sheet to their latest tags
# Excel has columns: 天賦, 姓名, 最新標籤
# We should clean the names in the Excel sheet because they might contain trailing spaces or newlines (e.g. '曾詩涵\n')
excel_data = {}
for r in range(2, sheet.max_row + 1):
    name_val = sheet.cell(row=r, column=2).value
    if name_val is not None:
        name = str(name_val).strip()
        tags_val = sheet.cell(row=r, column=3).value
        tags_str = str(tags_val) if tags_val is not None else ""
        
        # Split tags by newline
        raw_tags = [t.strip() for t in tags_str.split("\n") if t.strip()]
        # Limit to at most 3 tags
        limited_tags = raw_tags[:3]
        
        # If we have less than 3 tags, or if there's no tags, we keep it as list
        excel_data[name] = limited_tags

# 3. Match parents to the excel data
matched_parents = {}
unmatched_parents = []

for p_name in parent_names:
    if p_name in excel_data:
        matched_parents[p_name] = excel_data[p_name]
    else:
        # Try substring match
        matched_name = None
        for ex_name in excel_data.keys():
            if p_name in ex_name or ex_name in p_name:
                matched_name = ex_name
                break
        if matched_name:
            matched_parents[p_name] = excel_data[matched_name]
            print(f"Sub-match parent: '{p_name}' matches '{matched_name}'")
        else:
            unmatched_parents.append(p_name)

print(f"Successfully matched tags for {len(matched_parents)} / {len(parent_names)} parents.")
if unmatched_parents:
    print("WARNING: Unmatched parents:", unmatched_parents)

# Save matched parents tags to JSON for inspection
with open(r"d:\Antigravity\Rich\scratch\new_tags.json", "w", encoding="utf-8") as f:
    json.dump(matched_parents, f, ensure_ascii=False, indent=2)

print("Saved scratch/new_tags.json")
